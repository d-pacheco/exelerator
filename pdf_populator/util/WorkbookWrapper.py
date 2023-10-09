import openpyxl
from ..exceptions.PythonDataSheetNotFoundException import PythonDataSheetNotFoundException

class WorkbookWrapper:
    def __init__(self, wb: openpyxl.Workbook, data_sheet_name: str):
        self.wb = wb
        self.data = {}
        self.data_sheet_name = data_sheet_name
        if self.data_sheet_name not in self.wb.sheetnames:
            raise PythonDataSheetNotFoundException(self.data_sheet_name)
        self.LoadDataFromWorkbook()
        random = 3

    def LoadDataFromWorkbook(self) -> None:
        # Loads and stores the key:value pairs from the data sheet name in the excel file to be looked up later
        # during the pdf population steps
        data_sheet = self.wb[self.data_sheet_name]
        for row_index, row in enumerate(data_sheet.iter_rows(min_row=2, values_only=True), start=2):
            column1_value = row[0]
            column2_value = row[1]
            self.data[column1_value] = column2_value
        return

    def GetValueFromKey(self, key_to_look_up: str):
        key_list = [key for key in self.data]
        for key in key_list:
            if key in key_to_look_up:
                return self.data[key]
        print(f"Could not find key {key_to_look_up} within {self.data_sheet_name}")
        return None
