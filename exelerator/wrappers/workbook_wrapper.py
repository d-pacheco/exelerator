import openpyxl
import logging
from typing import Optional, Dict
from exelerator.exceptions.data_sheet_not_found_exception import DataSheetNotFoundException

logger = logging.getLogger("exelerator")


class WorkbookWrapper:
    def __init__(self, excel_file_path, data_sheet_name: str):
        self.wb = openpyxl.load_workbook(excel_file_path, data_only=True)
        self.data_sheet_name = data_sheet_name
        if self.data_sheet_name not in self.wb.sheetnames:
            raise DataSheetNotFoundException(self.data_sheet_name)
        self.data = load_data_from_workbook(self.wb[data_sheet_name])

    def get_value_from_key(self, key_to_look_up: str) -> Optional[str]:
        key_list = [key for key in self.data]
        for key in key_list:
            if key in key_to_look_up:
                return self.data[key]
        # print(f"Could not find key {key_to_look_up} within {self.data_sheet_name}")
        logger.warning(f"Could not find key {key_to_look_up} within {self.data_sheet_name}")
        return None


def load_data_from_workbook(data_sheet) -> Dict[str, Optional[str]]:
    # Loads and stores the key:value pairs from the data sheet name in the Excel file to be looked up later
    # during the pdf population steps
    data: Dict[str, Optional[str]] = {}
    for row_index, row in enumerate(data_sheet.iter_rows(min_row=2, values_only=True), start=2):
        column1_value = row[0]
        column2_value = row[1]
        data[column1_value] = column2_value
    return data
